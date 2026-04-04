from __future__ import annotations

import json
import re
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook


SUMO_FUNCTION_RULES = {
    "Msat": {
        "argc": 2,
        "builder": lambda a: f"({a[0]}/({a[1]}+{a[0]}))",
        "semantics": ["monod_saturation"],
    },
    "Minh": {
        "argc": 2,
        "builder": lambda a: f"({a[1]}/({a[1]}+{a[0]}))",
        "semantics": ["inhibition"],
    },
    "Bellinh": {
        "argc": 3,
        "builder": lambda a: (
            f"((1+2*(10**(0.5*({a[1]}-{a[2]}))))/"
            f"(1+(10**({a[0]}-{a[2]}))+(10**({a[1]}-{a[0]}))))"
        ),
        "semantics": ["logic_switch", "ph_inhibition"],
    },
    "Logsat": {
        "argc": 3,
        "builder": lambda a: f"(1/(1+exp(({a[2]}-{a[0]})/{a[1]})))",
        "semantics": ["logic_switch", "saturation_switch"],
    },
    "Loginh": {
        "argc": 3,
        "builder": lambda a: f"(1/(1+exp(({a[0]}-{a[2]})/{a[1]})))",
        "semantics": ["logic_switch", "inhibition_switch"],
    },
    "Lin": {
        "argc": 3,
        "builder": lambda a: f"({a[1]}*{a[0]}+{a[2]})",
        "semantics": ["linear"],
    },
    "Hardbound": {
        "argc": 3,
        "builder": lambda a: f"max({a[1]}; min({a[2]}; {a[0]}))",
        "semantics": ["bounds"],
    },
    "Arrh": {
        "argc": 3,
        "builder": lambda a: f"({a[0]}**({a[1]}-{a[2]}))",
        "semantics": ["temperature_correction"],
    },
    "MinuslogVar": {
        "argc": 1,
        "builder": lambda a: f"(10**(-{a[0]}))",
        "semantics": ["log_transform"],
    },
    "MRsat": {
        "argc": 3,
        "builder": lambda a: f"({a[0]}/({a[2]}*{a[1]}+{a[0]}))",
        "semantics": ["monod_saturation", "ratio_saturation"],
    },
    "MRinh": {
        "argc": 3,
        "builder": lambda a: f"({a[2]}*{a[1]}/({a[2]}*{a[1]}+{a[0]}))",
        "semantics": ["inhibition", "ratio_inhibition"],
    },
}

SUMO_FUNCTION_NAMES = "|".join(SUMO_FUNCTION_RULES.keys())
SUMO_EXPLICIT_PATTERN = re.compile(rf"(?<!\w)({SUMO_FUNCTION_NAMES})\(([^()]+)\)")
SUMO_COMPACT_PATTERN = re.compile(
    rf"(?<!\w)({SUMO_FUNCTION_NAMES})([A-Za-z0-9]+(?:[,_][A-Za-z0-9]+)+)(?!\w)"
)
IDENTIFIER_PATTERN = re.compile(r"[A-Za-z_][A-Za-z0-9_]*")
RAW_IDENTIFIER_PATTERN = re.compile(r"[A-Za-z_µμƞη][A-Za-z0-9_,µμƞη]*")
BACKEND_BUILTINS = {"exp", "log", "sqrt", "pow", "min", "max", "abs", "clip", "pi", "e"}
DEFAULT_EXCLUDED_COLUMNS = {"j", "Symbol", "Name", "Rate", "Unit", "Reaction", "Rule"}
PROCESS_NAME_DEFAULT = "symbol"


@dataclass
class AliasEntry:
    canonical_name: str
    sumo_name: str
    sumo_variants: List[str] = field(default_factory=list)
    unit: str = ""
    dimension_tags: List[str] = field(default_factory=list)
    conversion_factors: Dict[str, float] = field(default_factory=dict)
    hybrid_equivalence_group: str = ""


@dataclass
class NormalizedProcessRecord:
    process_id: str
    process_name: str
    sumo_symbol: str
    display_name: str
    rate_expr_raw: str
    rate_expr_platform: str
    rate_function_kinds: List[str]
    rate_semantics: List[str]
    note: str
    stoich_by_component: Dict[str, str]
    component_aliases: List[Dict[str, Any]]
    focal_vars: List[str]
    tags: Dict[str, List[str]]
    source_ref: Dict[str, Any]


@dataclass
class NormalizedProcessCatalog:
    source: str
    source_sheet: str
    process_name_mode: str
    component_alias_registry: List[AliasEntry]
    processes: List[NormalizedProcessRecord]
    canonical_components: List[str]


@dataclass
class SliceDefinition:
    name: str
    description: str
    include: Dict[str, List[str]] = field(default_factory=dict)
    exclude: Dict[str, List[str]] = field(default_factory=dict)
    component_mode: str = "active_only"


def normalize_sumo_args(parts: List[str], count: int) -> List[str]:
    if len(parts) < count:
        return []
    if len(parts) == count:
        return parts
    return parts[: count - 1] + ["_".join(parts[count - 1 :])]


def normalize_identifier(text: str) -> str:
    raw = str(text or "").strip()
    if raw == "":
        return ""
    raw = (
        raw.replace("μ", "miu")
        .replace("µ", "miu")
        .replace("η", "eta")
        .replace("ƞ", "eta")
        .replace("−", "-")
    )
    raw = re.sub(r"\s+", "", raw)
    return raw.replace(",", "_")


def slugify(text: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "_", str(text or "").strip().lower())
    return re.sub(r"_+", "_", value).strip("_")


def read_source_table(path: Path, sheet_name: str) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in [".xlsx", ".xlsm", ".xls"]:
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str)
    if suffix == ".csv":
        last_error: Optional[Exception] = None
        for enc in ["utf-8-sig", "utf-8", "gb18030", "gbk", "gb2312"]:
            try:
                return pd.read_csv(path, encoding=enc, dtype=str)
            except Exception as exc:  # pragma: no cover
                last_error = exc
        raise RuntimeError(f"无法读取 CSV: {path}") from last_error
    raise ValueError(f"不支持的输入格式: {path.suffix}")


def source_component_columns(df: pd.DataFrame) -> List[str]:
    return [column for column in df.columns if column not in DEFAULT_EXCLUDED_COLUMNS]


def dedupe_names(items: Sequence[str]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for item in items:
        base = item or "process"
        if base not in seen:
            seen[base] = 1
            result.append(base)
            continue
        seen[base] += 1
        result.append(f"{base}_{seen[base]}")
    return result


def unique_strings(items: Sequence[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for item in items:
        value = str(item or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def detect_rate_function_kinds(expr: str) -> List[str]:
    kinds = []
    for name in SUMO_FUNCTION_RULES:
        if re.search(rf"(?<!\w){name}(?:\(|[A-Za-z0-9])", expr):
            kinds.append(name)
    return kinds


def detect_rate_semantics(function_kinds: Sequence[str]) -> List[str]:
    seen = set()
    output: List[str] = []
    for kind in function_kinds:
        for semantic in SUMO_FUNCTION_RULES.get(kind, {}).get("semantics", []):
            if semantic in seen:
                continue
            seen.add(semantic)
            output.append(semantic)
    return output


def normalize_raw_identifiers(text: str) -> str:
    def repl(match: re.Match[str]) -> str:
        return normalize_identifier(match.group(0))

    return RAW_IDENTIFIER_PATTERN.sub(repl, text)


def expand_sumo_tokens(text: str) -> str:
    current = text

    def repl_explicit(match: re.Match[str]) -> str:
        kind = match.group(1)
        arg_raw = match.group(2)
        parts = [normalize_identifier(p) for p in re.split(r"[;,]", arg_raw) if p.strip()]
        rule = SUMO_FUNCTION_RULES[kind]
        args = normalize_sumo_args(parts, rule["argc"])
        if not args:
            return match.group(0)
        return rule["builder"](args)

    def repl_compact(match: re.Match[str]) -> str:
        kind = match.group(1)
        tail = match.group(2)
        parts = [normalize_identifier(p) for p in re.split(r"[,_]", tail) if p]
        rule = SUMO_FUNCTION_RULES[kind]
        args = normalize_sumo_args(parts, rule["argc"])
        if not args:
            return match.group(0)
        return rule["builder"](args)

    for _ in range(8):
        next_value = SUMO_EXPLICIT_PATTERN.sub(repl_explicit, current)
        next_value = SUMO_COMPACT_PATTERN.sub(repl_compact, next_value)
        if next_value == current:
            break
        current = next_value
    return current


def clean_expr(expr: Any) -> str:
    if expr is None:
        return ""
    raw = str(expr).strip()
    if raw == "" or raw.lower() == "nan":
        return ""
    normalized = (
        raw.replace("μ", "miu")
        .replace("µ", "miu")
        .replace("η", "eta")
        .replace("ƞ", "eta")
        .replace("−", "-")
    )
    normalized = expand_sumo_tokens(normalized)
    normalized = normalize_raw_identifiers(normalized)
    normalized = normalized.replace("^", "**")
    normalized = normalized.replace(";", ",")
    normalized = re.sub(r"\s*([+\-/(),])\s*", r" \1 ", normalized)
    normalized = re.sub(r"(?<!\*)\*(?!\*)", " * ", normalized)
    normalized = re.sub(r"\s*\*\*\s*", " ** ", normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def infer_component_dimensions(canonical_name: str) -> List[str]:
    name = canonical_name.upper()
    tags: List[str] = []

    nitrogen_markers = [
        "NH",
        "NO",
        "N2",
        "XN_",
        "SN_",
        "GNH3",
        "GN2",
        "GNO",
        "NOB",
        "AOB",
        "AMX",
    ]
    phosphorus_markers = [
        "PO4",
        "PP",
        "SP_",
        "XP_",
        "ACP",
        "BSH",
        "STR",
        "VIVI",
        "HFO",
        "HAO",
    ]
    alkalinity_markers = ["ALK", "CAT", "AN", "CA", "MG", "K", "INORG", "CO2"]
    carbon_markers = [
        "VFA",
        "SB",
        "SMEOL",
        "CB",
        "XB",
        "SU",
        "CU",
        "XU",
        "PHA",
        "XE",
        "OHO",
        "CASTO",
        "MEOLO",
        "AMETO",
        "HMETO",
        "ALGAE",
        "CH4",
        "H2",
        "CO2",
        "PFOA",
        "PFOS",
        "ALPHA",
    ]

    if any(marker in name for marker in nitrogen_markers):
        tags.append("nitrogen")
    if any(marker in name for marker in phosphorus_markers):
        tags.append("phosphorus")
    if canonical_name in {"S_ALK", "SCAT", "SAN", "SCa", "SMg", "SK", "SCO2", "XINORG"} or any(
        marker in name for marker in alkalinity_markers
    ):
        tags.append("alkalinity")
    if any(marker in name for marker in carbon_markers):
        tags.append("carbon")

    if not tags and canonical_name.startswith(("S", "X", "G")):
        tags.append("carbon")
    return unique_strings(tags)


def _load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(path)
    return json.loads(path.read_text(encoding="utf-8"))


def load_component_alias_registry(
    source_columns: Sequence[str],
    config_path: Optional[Path] = None,
) -> Tuple[Dict[str, AliasEntry], List[AliasEntry]]:
    config: Dict[str, Any] = {"aliases": []}
    if config_path and config_path.exists():
        config = _load_json(config_path)

    overrides: Dict[str, Dict[str, Any]] = {}
    for raw_entry in config.get("aliases", []):
        if not isinstance(raw_entry, dict):
            continue
        keys = [normalize_identifier(raw_entry.get("sumo_name", ""))]
        keys.extend(normalize_identifier(item) for item in raw_entry.get("sumo_variants", []) if item)
        for key in keys:
            if key:
                overrides[key] = raw_entry

    registry_map: Dict[str, AliasEntry] = {}
    registry_list: List[AliasEntry] = []
    for column in source_columns:
        normalized_column = normalize_identifier(column)
        override = overrides.get(normalized_column, {})
        entry = AliasEntry(
            canonical_name=normalize_identifier(override.get("canonical_name") or column),
            sumo_name=column,
            sumo_variants=[normalize_identifier(item) for item in override.get("sumo_variants", []) if item],
            unit=str(override.get("unit", "")).strip(),
            dimension_tags=override.get("dimension_tags")
            or infer_component_dimensions(normalize_identifier(override.get("canonical_name") or column)),
            conversion_factors=override.get("conversion_factors") or {},
            hybrid_equivalence_group=str(
                override.get("hybrid_equivalence_group")
                or normalize_identifier(override.get("canonical_name") or column)
            ).strip(),
        )
        registry_map[column] = entry
        registry_list.append(entry)
    return registry_map, registry_list


def build_process_names(df: pd.DataFrame, mode: str) -> List[str]:
    symbols = df["Symbol"].fillna("").astype(str).str.strip().tolist()
    names = df["Name"].fillna("").astype(str).str.strip().tolist()
    raw: List[str] = []
    for symbol, name in zip(symbols, names):
        if mode == "name":
            raw.append(name or symbol or "process")
        elif mode == "name_slug":
            raw.append(slugify(name) or symbol or "process")
        else:
            raw.append(symbol or slugify(name) or "process")
    normalized = [normalize_identifier(item) if mode != "name" else item for item in raw]
    return dedupe_names(normalized)


def extract_focal_variables(rate_expr: str, component_names: Iterable[str]) -> List[str]:
    component_set = set(component_names)
    result: List[str] = []
    seen = set()
    for token in IDENTIFIER_PATTERN.findall(rate_expr):
        if token in BACKEND_BUILTINS or token not in component_set or token in seen:
            continue
        seen.add(token)
        result.append(token)
    return result


def infer_guilds(display_name: str, involved_components: Sequence[str]) -> List[str]:
    text = f"{display_name} {' '.join(involved_components)}".upper()
    tags: List[str] = []

    if "OHO" in text or "HETEROTROPH" in text:
        tags.append("heterotroph")
    if "AOB" in text:
        tags.extend(["aob", "autotroph"])
    if "NOB" in text:
        tags.extend(["nob", "autotroph"])
    if "AMX" in text:
        tags.extend(["amx", "autotroph"])
    if "CASTO" in text or "PAO" in text:
        tags.append("pao")
    if "GAO" in text:
        tags.append("gao")
    if "ALGAE" in text:
        tags.append("autotroph")
    if "ABIOTIC" in text or any(
        keyword in text
        for keyword in ["PRECIPITATION", "TRANSFER", "SORPTION", "DESORP", "DISSOLUTION", "OXIDATION OF FE2"]
    ):
        tags.append("abiotic")

    if not tags and any(token in text for token in ["AMETO", "HMETO", "MEOLO"]):
        tags.append("heterotroph")
    return unique_strings(tags or ["general"])


def infer_redox_modes(display_name: str, rate_expr: str) -> List[str]:
    upper_name = display_name.upper()
    upper_expr = rate_expr.upper()
    tags: List[str] = []
    if "O2" in upper_name or "AEROBIC" in upper_name or "SO2" in upper_expr:
        tags.append("aerobic")
    if "NO3" in upper_name or "SNO3" in upper_expr:
        tags.append("anoxic_no3")
    if "NO2" in upper_name or "SNO2" in upper_expr:
        tags.append("anoxic_no2")
    if re.search(r"(?<!N)NO(?!2|3)", upper_name) or "SNO_AOB" in upper_expr or "SNO_OHO" in upper_expr:
        tags.append("anoxic_no")
    if "N2O" in upper_name or "SN2O" in upper_expr:
        tags.append("anoxic_n2o")
    if "ANAEROBIC" in upper_name or "ANAEROBIC" in upper_expr:
        tags.append("anaerobic")
    return unique_strings(tags or ["general"])


def infer_lifecycle(display_name: str) -> List[str]:
    upper_name = display_name.upper()
    mappings = {
        "growth": ["GROWTH"],
        "decay": ["DECAY", "RESPIRATION"],
        "hydrolysis": ["HYDROLYSIS"],
        "storage": ["STORAGE", "ASSIMILATIVE"],
        "release": ["RELEASE", "DESORP", "DISSOLUTION", "CLEAVAGE"],
        "transfer": [
            "TRANSFER",
            "SORPTION",
            "OXIDATION",
            "REDUCTION",
            "FERMENTATION",
            "CONVERSION",
            "MAINTENANCE",
        ],
        "precipitation": ["PRECIPITATION", "BINDING", "COPREC"],
    }
    tags: List[str] = []
    for tag, keywords in mappings.items():
        if any(keyword in upper_name for keyword in keywords):
            tags.append(tag)
    return unique_strings(tags or ["transfer"])


def infer_domain(display_name: str, involved_components: Sequence[str], guilds: Sequence[str]) -> List[str]:
    upper_text = f"{display_name} {' '.join(involved_components)}".upper()
    if any(token in upper_text for token in ["DIGEST", "CH4", "GH2", "GCH4", "SCH4", "SH2", "AMETO", "HMETO"]):
        return ["digestion"]
    if "SIDESTREAM" in upper_text:
        return ["sidestream"]
    if "abiotic" in guilds:
        return ["general"]
    return ["activated_sludge"]


def infer_extension_flags(
    display_name: str,
    involved_components: Sequence[str],
    guilds: Sequence[str],
    redox_modes: Sequence[str],
    element_scope: Sequence[str],
) -> List[str]:
    upper_text = f"{display_name} {' '.join(involved_components)}".upper()
    flags: List[str] = []
    if any(tag in guilds for tag in ["aob", "nob"]) or any(
        token in upper_text for token in ["NH2OH", "AOB", "NOB", "NITRIFICATION"]
    ):
        flags.append("two_step_nitrification")
    if any(token in upper_text for token in ["AMETO", "HMETO", "CH4", "H2", "DIGEST"]):
        flags.append("two_step_digestion")
    if any(mode in redox_modes for mode in ["anoxic_no3", "anoxic_no2", "anoxic_no", "anoxic_n2o"]) or any(
        token in upper_text for token in ["NO3", "NO2", " N2O", " NO "]
    ):
        flags.append("four_step_n_conversion")
    if any(
        token in upper_text
        for token in ["PAO", "GAO", "POLYPHOSPHATE", "XPP", "XPHA", "HFO", "HAO", "STR", "VIVI", "ACP", "BSH"]
    ):
        flags.append("p_removal")
    return unique_strings(flags)


def infer_process_tags(
    *,
    display_name: str,
    rate_expr_platform: str,
    component_aliases: Sequence[Dict[str, Any]],
) -> Dict[str, List[str]]:
    involved_components = [str(item.get("canonical_name", "")) for item in component_aliases]
    dimension_tags = [
        tag
        for item in component_aliases
        for tag in item.get("dimension_tags", [])
        if isinstance(tag, str) and tag
    ]
    element_scope = unique_strings(dimension_tags or ["carbon"])
    guilds = infer_guilds(display_name, involved_components)
    redox_modes = infer_redox_modes(display_name, rate_expr_platform)
    lifecycle = infer_lifecycle(display_name)
    domain = infer_domain(display_name, involved_components, guilds)
    extension_flags = infer_extension_flags(display_name, involved_components, guilds, redox_modes, element_scope)
    return {
        "guild": guilds,
        "element_scope": element_scope,
        "redox_mode": redox_modes,
        "lifecycle": lifecycle,
        "domain": domain,
        "extension_flags": extension_flags,
    }


def build_process_catalog(
    df: pd.DataFrame,
    *,
    source_path: Path,
    source_sheet: str,
    process_name_mode: str = PROCESS_NAME_DEFAULT,
    alias_config_path: Optional[Path] = None,
) -> NormalizedProcessCatalog:
    required = ["Symbol", "Name", "Rate"]
    missing_required = [column for column in required if column not in df.columns]
    if missing_required:
        raise ValueError(f"源数据缺少必要列: {missing_required}")

    work = df.copy()
    work["Symbol"] = work["Symbol"].fillna("").astype(str).str.strip()
    work["Name"] = work["Name"].fillna("").astype(str).str.strip()
    work["Rate"] = work["Rate"].fillna("").astype(str).str.strip()
    work = work[(work["Symbol"] != "") & (work["Rate"] != "")]

    component_columns = source_component_columns(work)
    alias_by_source, alias_registry = load_component_alias_registry(component_columns, alias_config_path)
    canonical_components = dedupe_names([entry.canonical_name for entry in alias_registry])
    process_names = build_process_names(work, process_name_mode)

    records: List[NormalizedProcessRecord] = []
    for index, (_, row) in enumerate(work.iterrows()):
        rate_expr_raw = str(row.get("Rate", "")).strip()
        rate_expr_platform = clean_expr(rate_expr_raw)
        rate_function_kinds = detect_rate_function_kinds(rate_expr_raw)
        rate_semantics = detect_rate_semantics(rate_function_kinds)

        stoich_by_component: Dict[str, str] = {}
        alias_rows: List[Dict[str, Any]] = []
        for component_column in component_columns:
            raw_value = clean_expr(row.get(component_column, ""))
            if raw_value == "" or raw_value == "0":
                continue
            alias = alias_by_source[component_column]
            stoich_by_component[alias.canonical_name] = raw_value
            alias_rows.append(
                {
                    "canonical_name": alias.canonical_name,
                    "sumo_name": alias.sumo_name,
                    "dimension_tags": alias.dimension_tags,
                    "hybrid_equivalence_group": alias.hybrid_equivalence_group,
                }
            )

        display_name = str(row.get("Name", "")).strip()
        note_parts = [str(row.get("Unit", "")).strip(), display_name]
        note = " ".join(part for part in note_parts if part).strip()
        focal_vars = extract_focal_variables(rate_expr_platform, canonical_components)
        tags = infer_process_tags(
            display_name=display_name,
            rate_expr_platform=rate_expr_platform,
            component_aliases=alias_rows,
        )
        process_name = process_names[index]
        process_id = slugify(f"{row['Symbol']}_{display_name}") or process_name
        records.append(
            NormalizedProcessRecord(
                process_id=process_id,
                process_name=process_name,
                sumo_symbol=str(row.get("Symbol", "")).strip(),
                display_name=display_name,
                rate_expr_raw=rate_expr_raw,
                rate_expr_platform=rate_expr_platform,
                rate_function_kinds=rate_function_kinds,
                rate_semantics=rate_semantics,
                note=note,
                stoich_by_component=stoich_by_component,
                component_aliases=alias_rows,
                focal_vars=focal_vars,
                tags=tags,
                source_ref={
                    "source": str(source_path),
                    "sheet": source_sheet,
                    "row_index": index,
                    "excel_row": index + 2,
                },
            )
        )

    return NormalizedProcessCatalog(
        source=str(source_path),
        source_sheet=source_sheet,
        process_name_mode=process_name_mode,
        component_alias_registry=alias_registry,
        processes=records,
        canonical_components=canonical_components,
    )


def load_decomposition_profile(config_path: Optional[Path]) -> List[SliceDefinition]:
    if config_path is None or not config_path.exists():
        return []
    payload = _load_json(config_path)
    slices: List[SliceDefinition] = []
    for raw in payload.get("slices", []):
        if not isinstance(raw, dict):
            continue
        slices.append(
            SliceDefinition(
                name=str(raw.get("name", "")).strip(),
                description=str(raw.get("description", "")).strip(),
                include={
                    key: [str(item).strip() for item in value if str(item).strip()]
                    for key, value in (raw.get("include") or {}).items()
                    if isinstance(value, list)
                },
                exclude={
                    key: [str(item).strip() for item in value if str(item).strip()]
                    for key, value in (raw.get("exclude") or {}).items()
                    if isinstance(value, list)
                },
                component_mode=str(raw.get("component_mode") or "active_only").strip(),
            )
        )
    return [item for item in slices if item.name]


def _tags_match_axis(record: NormalizedProcessRecord, axis: str, values: Sequence[str]) -> bool:
    record_values = set(record.tags.get(axis, []))
    return bool(record_values.intersection(values))


def record_matches_slice(record: NormalizedProcessRecord, definition: SliceDefinition) -> bool:
    for axis, values in definition.include.items():
        if not _tags_match_axis(record, axis, values):
            return False
    for axis, values in definition.exclude.items():
        if _tags_match_axis(record, axis, values):
            return False
    return True


def build_slice_component_order(
    records: Sequence[NormalizedProcessRecord],
    registry: Sequence[AliasEntry],
    *,
    component_mode: str,
    full_replica: bool,
) -> List[str]:
    ordered_all = dedupe_names([entry.canonical_name for entry in registry])
    if full_replica or component_mode == "full_replica":
        return ordered_all
    active = {component for record in records for component in record.stoich_by_component}
    active.update(component for record in records for component in record.focal_vars)
    return [component for component in ordered_all if component in active]


def build_matrix_df(records: Sequence[NormalizedProcessRecord], component_order: Sequence[str]) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for record in records:
        row = {
            "process_name": record.process_name,
            "rate_expr": record.rate_expr_platform,
            "note": record.note,
        }
        for component in component_order:
            row[component] = record.stoich_by_component.get(component, "0")
        rows.append(row)
    return pd.DataFrame(rows, columns=["process_name", "rate_expr", "note", *component_order])


def write_to_template(
    template_path: Path,
    output_path: Path,
    matrix_df: pd.DataFrame,
    matrix_sheet_name: str,
    rewrite_header: bool,
) -> None:
    workbook = load_workbook(template_path)
    if matrix_sheet_name not in workbook.sheetnames:
        raise ValueError(f"模板中不存在工作表: {matrix_sheet_name}")
    worksheet = workbook[matrix_sheet_name]
    header = [worksheet.cell(1, i).value for i in range(1, worksheet.max_column + 1)]
    expected = matrix_df.columns.tolist()
    if not rewrite_header and header[: len(expected)] != expected:
        raise ValueError("模板表头与转换结果列不一致，请确认模板第一行列名顺序")
    if rewrite_header:
        if worksheet.max_column < len(expected):
            worksheet.insert_cols(worksheet.max_column + 1, len(expected) - worksheet.max_column)
        for i, name in enumerate(expected, start=1):
            worksheet.cell(1, i).value = name
        for i in range(len(expected) + 1, worksheet.max_column + 1):
            worksheet.cell(1, i).value = None
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for _, row in matrix_df.iterrows():
        worksheet.append([str(row[col]) if row[col] is not None else "" for col in expected])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_new_workbook(output_path: Path, matrix_df: pd.DataFrame, matrix_sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = matrix_sheet_name
    worksheet.append(matrix_df.columns.tolist())
    for _, row in matrix_df.iterrows():
        worksheet.append([str(row[col]) if row[col] is not None else "" for col in matrix_df.columns])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _load_backend_validator() -> Optional[Any]:
    project_root = Path(__file__).resolve().parents[2]
    backend_root = project_root / "backend"
    if not backend_root.exists():
        return None
    if str(backend_root) not in sys.path:
        sys.path.insert(0, str(backend_root))
    try:
        from app.services.udm_expression import validate_udm_definition
    except Exception:
        return None
    return validate_udm_definition


def validate_records_against_platform(
    records: Sequence[NormalizedProcessRecord],
    component_order: Sequence[str],
) -> Dict[str, Any]:
    validator = _load_backend_validator()
    if validator is None:
        return {"status": "skipped", "errors": [], "warnings": [], "extracted_parameters": []}

    result = validator(
        components=list(component_order),
        processes=[
            {
                "name": record.process_name,
                "rate_expr": record.rate_expr_platform,
                "stoich_expr": record.stoich_by_component,
            }
            for record in records
        ],
        declared_parameters=[],
    )
    return {
        "status": "ok" if result.ok else "error",
        "errors": [asdict(item) for item in result.errors],
        "warnings": [asdict(item) for item in result.warnings],
        "extracted_parameters": result.extracted_parameters,
    }


def catalog_to_dict(catalog: NormalizedProcessCatalog) -> Dict[str, Any]:
    return {
        "source": catalog.source,
        "source_sheet": catalog.source_sheet,
        "process_name_mode": catalog.process_name_mode,
        "canonical_components": catalog.canonical_components,
        "component_alias_registry": [asdict(item) for item in catalog.component_alias_registry],
        "processes": [asdict(item) for item in catalog.processes],
    }


def write_catalog_json(catalog: NormalizedProcessCatalog, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(catalog_to_dict(catalog), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def generate_catalog_report(catalog: NormalizedProcessCatalog) -> str:
    lines = [
        "# SUMO Petersen Catalog Report",
        "",
        f"- Source: `{catalog.source}`",
        f"- Sheet: `{catalog.source_sheet}`",
        f"- Processes: `{len(catalog.processes)}`",
        f"- Canonical components: `{len(catalog.canonical_components)}`",
        "",
        "## Tag Summary",
        "",
    ]
    for axis in ["guild", "element_scope", "redox_mode", "lifecycle", "domain", "extension_flags"]:
        counts: Dict[str, int] = {}
        for record in catalog.processes:
            for value in record.tags.get(axis, []):
                counts[value] = counts.get(value, 0) + 1
        lines.append(f"### {axis}")
        if not counts:
            lines.append("- None")
            lines.append("")
            continue
        for key in sorted(counts):
            lines.append(f"- `{key}`: {counts[key]}")
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def generate_slice_report(
    definition: SliceDefinition,
    records: Sequence[NormalizedProcessRecord],
    component_order: Sequence[str],
    validation: Dict[str, Any],
) -> str:
    lines = [
        f"# Slice Report: {definition.name}",
        "",
        definition.description,
        "",
        f"- Processes: `{len(records)}`",
        f"- Components: `{len(component_order)}`",
        f"- Platform validation: `{validation['status']}`",
        "",
        "## Components",
        "",
    ]
    for component in component_order:
        lines.append(f"- `{component}`")
    lines.extend(["", "## Processes", ""])
    for record in records:
        tag_summary = ", ".join(f"{axis}={'+'.join(values)}" for axis, values in record.tags.items() if values)
        lines.append(
            f"- `{record.process_name}` ({record.sumo_symbol})"
            f": {record.display_name} | semantics={','.join(record.rate_semantics) or 'none'}"
        )
        lines.append(f"  - Tags: {tag_summary or 'none'}")
    if validation["errors"]:
        lines.extend(["", "## Validation Errors", ""])
        for item in validation["errors"]:
            lines.append(f"- `{item['code']}`: {item['message']}")
    if validation["warnings"]:
        lines.extend(["", "## Validation Warnings", ""])
        for item in validation["warnings"]:
            lines.append(f"- `{item['code']}`: {item['message']}")
    return "\n".join(lines).strip() + "\n"


def export_slice_workbook(
    *,
    records: Sequence[NormalizedProcessRecord],
    registry: Sequence[AliasEntry],
    output_path: Path,
    template_path: Optional[Path],
    template_sheet: str,
    component_mode: str,
    full_replica: bool,
) -> Tuple[pd.DataFrame, List[str]]:
    component_order = build_slice_component_order(
        records,
        registry,
        component_mode=component_mode,
        full_replica=full_replica,
    )
    matrix_df = build_matrix_df(records, component_order)
    if template_path:
        write_to_template(
            template_path,
            output_path,
            matrix_df,
            template_sheet,
            rewrite_header=full_replica,
        )
    else:
        write_new_workbook(output_path, matrix_df, template_sheet)
    return matrix_df, component_order


def run_pipeline_exports(
    *,
    catalog: NormalizedProcessCatalog,
    slice_definitions: Sequence[SliceDefinition],
    output_dir: Path,
    template_path: Optional[Path],
    template_sheet: str,
    emit_catalog: bool,
    emit_markdown: bool,
    full_replica: bool,
) -> Dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    slices_dir = output_dir / "slices"
    reports_dir = output_dir / "reports"
    slices_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)

    if emit_catalog:
        write_catalog_json(catalog, output_dir / "catalog.json")
    if emit_markdown:
        (reports_dir / "catalog.md").write_text(generate_catalog_report(catalog), encoding="utf-8")

    manifest: Dict[str, Any] = {
        "source": catalog.source,
        "source_sheet": catalog.source_sheet,
        "slices": [],
    }
    for definition in slice_definitions:
        matched_records = [record for record in catalog.processes if record_matches_slice(record, definition)]
        slice_path = slices_dir / f"{definition.name}.xlsx"
        _, component_order = export_slice_workbook(
            records=matched_records,
            registry=catalog.component_alias_registry,
            output_path=slice_path,
            template_path=template_path,
            template_sheet=template_sheet,
            component_mode=definition.component_mode,
            full_replica=full_replica,
        )
        validation = validate_records_against_platform(matched_records, component_order)
        manifest["slices"].append(
            {
                "name": definition.name,
                "description": definition.description,
                "process_count": len(matched_records),
                "component_count": len(component_order),
                "components": component_order,
                "output": str(slice_path),
                "validation": validation,
            }
        )
        if emit_markdown:
            (reports_dir / f"{definition.name}.md").write_text(
                generate_slice_report(definition, matched_records, component_order, validation),
                encoding="utf-8",
            )

    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return manifest


def resolve_default_path(filename: str) -> Path:
    return Path(__file__).resolve().parent / filename
